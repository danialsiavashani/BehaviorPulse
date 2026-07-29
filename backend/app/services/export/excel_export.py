import io
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.worksheet.worksheet import Worksheet

from app.db.models.observation_analysis import ObservationAnalysis

TITLE_FONT = Font(bold=True, size=14)
LABEL_FONT = Font(bold=True)
HEADER_FONT = Font(bold=True)
WRAP = Alignment(wrap_text=True, vertical="top")


def build_analysis_export(analysis: ObservationAnalysis, app_name: str) -> bytes:
    """Builds an .xlsx workbook from an already-computed ObservationAnalysis
    row. This is pure formatting of existing, stored data - no new
    computation happens here. Tier 1 export only; raw observations are not
    persisted anywhere, so they can't be included (see Tier 2, deferred).
    """
    wb = Workbook()

    _build_summary_sheet(wb.active, analysis, app_name)
    _build_metrics_sheet(wb.create_sheet("Metrics"), analysis)
    _build_recommendations_sheet(wb.create_sheet("Recommendations"), analysis)

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def _build_summary_sheet(ws: Worksheet, analysis: ObservationAnalysis, app_name: str) -> None:
    ws.title = "Summary"
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 70

    ws["A1"] = "BehaviorPulse Analysis Export"
    ws["A1"].font = TITLE_FONT

    rows = [
        ("Analysis ID", analysis.analysis_id),
        ("App", app_name),
        ("Subject Type", analysis.subject_type),
        ("Subject Label", analysis.subject_label),
        ("Total Observations", analysis.total_observations),
        ("Prediction Confidence", analysis.computed_confidence),
        ("Created At", analysis.created_at.strftime("%Y-%m-%d %H:%M UTC") if analysis.created_at else ""),
    ]
    r = 3
    for label, value in rows:
        ws.cell(row=r, column=1, value=label).font = LABEL_FONT
        ws.cell(row=r, column=2, value=value)
        r += 1

    r += 1
    ws.cell(row=r, column=1, value="Summary").font = LABEL_FONT
    r += 1
    cell = ws.cell(row=r, column=1, value=analysis.summary)
    cell.alignment = WRAP
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
    ws.row_dimensions[r].height = 60

    r += 2
    ws.cell(row=r, column=1, value="Prediction").font = LABEL_FONT
    r += 1
    cell = ws.cell(row=r, column=1, value=analysis.prediction)
    cell.alignment = WRAP
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
    ws.row_dimensions[r].height = 60


def _write_table(ws: Worksheet, start_row: int, headers: list[str], rows: list[list]) -> int:
    """Writes a small header+rows table starting at start_row. Returns the
    next free row after the table (with one blank row spacing)."""
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=start_row, column=col, value=header)
        cell.font = HEADER_FONT
    for i, row in enumerate(rows, start=1):
        for col, value in enumerate(row, start=1):
            ws.cell(row=start_row + i, column=col, value=value)
    return start_row + len(rows) + 2


def _build_metrics_sheet(ws: Worksheet, analysis: ObservationAnalysis) -> None:
    metrics = analysis.computed_metrics_json
    for col in ("A", "B", "C"):
        ws.column_dimensions[col].width = 22

    r = 1
    ws.cell(row=r, column=1, value="Average Confidence").font = LABEL_FONT
    ws.cell(row=r, column=2, value=metrics.get("average_confidence"))
    r += 2

    ws.cell(row=r, column=1, value="Top Subjects").font = TITLE_FONT
    r += 1
    top_subjects = metrics.get("top_subjects") or []
    r = _write_table(
        ws, r,
        ["Subject", "Count", "Percentage"],
        [[s["subject_label"], s["count"], s["percentage"]] for s in top_subjects],
    )

    ws.cell(row=r, column=1, value="Top Sources").font = TITLE_FONT
    r += 1
    top_sources = metrics.get("top_sources") or []
    r = _write_table(
        ws, r,
        ["Source", "Count", "Percentage"],
        [[s["source_id"], s["count"], s["percentage"]] for s in top_sources],
    )

    ws.cell(row=r, column=1, value="Top Day of Week").font = TITLE_FONT
    r += 1
    top_day = metrics.get("top_day_of_week")
    if top_day:
        r = _write_table(ws, r, ["Day", "Count", "Percentage"], [[top_day["day"], top_day["count"], top_day["percentage"]]])
    else:
        ws.cell(row=r, column=1, value="No dominant day of week identified.")
        r += 2

    ws.cell(row=r, column=1, value="Top Time Window").font = TITLE_FONT
    r += 1
    top_window = metrics.get("top_time_window")
    if top_window:
        r = _write_table(ws, r, ["Window", "Count", "Percentage"], [[top_window["window"], top_window["count"], top_window["percentage"]]])
    else:
        ws.cell(row=r, column=1, value="No dominant time window identified.")
        r += 2

    ws.cell(row=r, column=1, value="Pattern Table").font = TITLE_FONT
    r += 1
    pattern_rows = analysis.pattern_table_json or []
    _write_table(
        ws, r,
        ["Metric", "Value", "Support"],
        [[row["metric"], row["value"], row["support"]] for row in pattern_rows],
    )
    ws.column_dimensions["C"].width = 50


def _build_recommendations_sheet(ws: Worksheet, analysis: ObservationAnalysis) -> None:
    ws.column_dimensions["A"].width = 90

    ws.cell(row=1, column=1, value="Recommendations").font = TITLE_FONT
    r = 2
    for rec in analysis.recommendations_json or []:
        cell = ws.cell(row=r, column=1, value=f"• {rec}")
        cell.alignment = WRAP
        r += 1

    r += 1
    ws.cell(row=r, column=1, value="Warnings").font = TITLE_FONT
    r += 1
    for warning in analysis.warnings_json or []:
        cell = ws.cell(row=r, column=1, value=f"• {warning}")
        cell.alignment = WRAP
        r += 1